import pandas as pd
import unicodedata
import re
import chardet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook

# --- Funções de limpeza e padronização ---

def remover_acentos(texto):
    if pd.isna(texto):
        return texto
    texto = str(texto)
    texto = unicodedata.normalize("NFD", texto)
    texto = re.sub(r"[\u0300-\u036f]", "", texto)
    return texto

def padronizar_texto_para_sgbd(texto):
    if pd.isna(texto):
        return texto
    s = str(texto)
    # substituir aspas e caracteres problemáticos
    s = s.replace(r'\u201c', '"').replace(r'\u201d', '"').replace("’", "'").replace("‘", "'")
    s = re.sub(r"[\r\n\t]+", " ", s)  # remover quebras de linha e tabs
    s = re.sub(r"\s+", " ", s)        # remover múltiplos espaços
    s = s.strip()
    s = remover_acentos(s)             # remover acentos
    s = s.upper()                      # padronizar maiúsculas
    if s in ['NAN', 'NONE', 'NA']:
        return pd.NA
    return s

def try_parse_datetime(col_series, dayfirst=True):
    parsed = pd.to_datetime(col_series, dayfirst=dayfirst, errors='coerce', infer_datetime_format=True)
    success_rate = parsed.notna().sum() / len(parsed)
    return parsed, success_rate

def normalize_numeric_string(s):
    if pd.isna(s):
        return s
    s = str(s).strip()
    if s.endswith('%'):
        try:
            return float(s.replace('%','').replace('.','').replace(',','.'))/100.0
        except:
            return s
    if re.search(r'[A-Za-z]', s):
        return s
    if '.' in s and ',' in s:
        s2 = s.replace('.', '').replace(',', '.')
        return s2
    if ',' in s and '.' not in s:
        s2 = s.replace('.', '').replace(',', '.')
        return s2
    s = s.replace(' ', '')
    return s

def detect_separator(file_path, encoding):
    with open(file_path, 'r', encoding=encoding, errors='replace') as f:
        first_line = f.readline(2000)
    for s in [',',';','\t','|']:
        if first_line.count(s) > 0:
            return s
    return ','

# --- Função para formatar Excel ---

def format_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # cabeçalhos em negrito e centralizados
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ajustar largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # centralizar células de texto
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str):
                cell.alignment = Alignment(horizontal='left', vertical='center')

    wb.save(file_path)
    print(f"Arquivo Excel formatado: {file_path}")

# --- Função principal ---

def main(input_csv='detalhe_votacao.csv', out_xlsx='dados_tratados.xlsx', out_csv='dados_tratados.csv'):
    # detectar encoding
    with open(input_csv, 'rb') as f:
        raw = f.read(500000)
        enc = chardet.detect(raw)['encoding'] or 'utf-8'
    print('Encoding detectado:', enc)

    sep = detect_separator(input_csv, enc)
    print('Delimitador detectado:', repr(sep))

    df = pd.read_csv(input_csv, encoding=enc, sep=sep, on_bad_lines='skip', low_memory=False)
    print('Dimensão original:', df.shape)

    # detectar colunas object
    obj_cols = df.select_dtypes(include=['object']).columns.tolist()

    # detectar e converter colunas de data
    date_candidates = [c for c in df.columns if re.search("(data|date|dt|hora|timestamp|dia|momento|nasc|nascimento)", c, re.IGNORECASE)]
    date_cols_detected = []
    for col in obj_cols:
        parsed, rate = try_parse_datetime(df[col])
        if rate >= 0.5 or col in date_candidates:
            parsed2, rate2 = try_parse_datetime(df[col], dayfirst=True)
            if rate2 < 0.2:
                parsed2b, rate2b = try_parse_datetime(df[col], dayfirst=False)
                if rate2b > rate2:
                    parsed2, rate2 = parsed2b, rate2b
            if rate2 >= 0.05:
                df[col] = parsed2
                date_cols_detected.append(col)

    # aplicar padronização textual nas colunas restantes
    obj_cols_after = df.select_dtypes(include=['object']).columns.tolist()
    text_cols_to_clean = [c for c in obj_cols_after if c not in date_cols_detected]
    for col in text_cols_to_clean:
        df[col] = df[col].apply(padronizar_texto_para_sgbd)

    # converter colunas numéricas escritas como texto
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]) or pd.api.types.is_datetime64_any_dtype(df[col]):
            continue
        sample = df[col].dropna().astype(str).head(1000)
        if sample.shape[0] == 0:
            continue
        normalized = sample.apply(normalize_numeric_string)
        coerced = pd.to_numeric(normalized, errors='coerce')
        conv_rate = coerced.notna().sum() / sample.shape[0]
        if conv_rate >= 0.6:
            df[col] = df[col].apply(normalize_numeric_string)
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # padronizar datas para ISO
    final_date_cols = [c for c in df.columns if pd.api.types.is_datetime64_any_dtype(df[c])]
    for c in final_date_cols:
        has_time = df[c].dropna().apply(lambda x: (x.hour != 0 or x.minute != 0 or x.second != 0)).any()
        if has_time:
            df[c] = df[c].dt.strftime('%Y-%m-%d %H:%M:%S')
        else:
            df[c] = df[c].dt.strftime('%Y-%m-%d')

    # remover linhas vazias e duplicatas
    df.dropna(how='all', inplace=True)
    df.drop_duplicates(inplace=True)

    # exportar
    df.to_excel(out_xlsx, index=False)
    df.to_csv(out_csv, index=False, encoding='utf-8-sig')

    # aplicar formatação no Excel
    format_excel(out_xlsx)

    print('Tratamento finalizado. Arquivos gerados:', out_xlsx, out_csv)

if __name__ == '__main__':
    main()
